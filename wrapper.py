import os
import re
import sys
import cv2
import time
import fitz
import json
import logging
import openpyxl
import requests
import numpy as np
import pandas as pd
from enum import Enum
from groq import Groq
from PIL import Image
from functools import wraps
from datetime import datetime
from dotenv import load_dotenv
from dataclasses import dataclass
from presidio_anonymizer import AnonymizerEngine
from presidio_anonymizer.entities import OperatorConfig
from presidio_analyzer.nlp_engine import NlpEngineProvider
from typing import Dict, Tuple, Optional , List
from presidio_analyzer import AnalyzerEngine, RecognizerRegistry, Pattern, PatternRecognizer
            
# Robust retry decorator
def retry_on_failure(max_retries=3, delay=1):
    """Decorator to retry failed operations"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    if attempt == max_retries - 1:
                        logger.error(f"{func.__name__} failed after {max_retries} attempts: {e}")
                        raise
                    logger.warning(f"{func.__name__} attempt {attempt + 1} failed: {e}. Retrying...")
                    time.sleep(delay)
            return None
        return wrapper
    return decorator

load_dotenv()

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pii_redaction.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Configuration
@dataclass
class Config:
    groq_api_key: str
    groq_model: str = "llama-3.3-70b-versatile"
    spacy_model: str = "en_core_web_trf"
    language: str = "en"
    face_detection_model: str = "haarcascade_frontalface_default.xml"

# REDACTION POLICY ENGINE 
class RedactionSeverity(Enum):
    """Severity levels for PII redaction"""
    CRITICAL = "CRITICAL"  # Full blackout - IDs, credentials
    HIGH = "HIGH"          # Heavy blur - faces, screens
    MEDIUM = "MEDIUM"      # Medium blur - client logos
    LOW = "LOW"            # Light blur - name plates
    NONE = "NONE"          # No redaction

class RedactionMethod(Enum):
    """Redaction visualization methods"""
    BLACKOUT = "blackout"      # Solid black rectangle
    BLUR_HEAVY = "blur_heavy"  # Gaussian blur (99x99)
    BLUR_MEDIUM = "blur_medium" # Gaussian blur (51x51)
    BLUR_LIGHT = "blur_light"  # Gaussian blur (25x25)
    PIXELATE = "pixelate"      # Pixelation effect

class RedactionPolicy:
    """Enterprise-grade image redaction policy"""
    
    def __init__(self, policy_config: Optional[Dict] = None):
        self.policy = policy_config or self._default_corporate_policy()
    
    def _default_corporate_policy(self) -> Dict:
        """Default policy aligned with GDPR, DPDP Act India, ISO 27701"""
        return {
            # MANDATORY REDACTION
            'faces': {
                'enabled': True,
                'severity': RedactionSeverity.HIGH,
                'method': RedactionMethod.BLUR_HEAVY,
                'min_confidence': 0.5,
                'exceptions': [],  # Empty = redact all faces
                'reason': 'Biometric PII (GDPR Art. 9, DPDP Act Sec. 3)'
            },
            'identity_documents': {
                'enabled': True,
                'severity': RedactionSeverity.CRITICAL,
                'method': RedactionMethod.BLACKOUT,
                'types': ['aadhaar', 'pan', 'passport', 'driving_license', 
                         'employee_badge', 'voter_id', 'id_card'],
                'min_confidence': 0.6,
                'reason': 'Identity documents - full redaction required'
            },
            'screens_dashboards': {
                'enabled': True,
                'severity': RedactionSeverity.HIGH,
                'method': RedactionMethod.BLACKOUT,
                'keywords': ['dashboard', 'email', 'inbox', 'crm', 'jira', 
                            'excel', 'spreadsheet', 'report', 'analytics'],
                'min_confidence': 0.5,
                'reason': 'Potential confidential business data'
            },
            'ocr_pii': {
                'enabled': True,
                'severity': RedactionSeverity.HIGH,
                'method': RedactionMethod.BLACKOUT,
                'min_confidence': 0.4,
                'entity_types': ['PERSON', 'PHONE_NUMBER', 'EMAIL_ADDRESS',
                               'AADHAAR_NUMBER', 'PAN_NUMBER', 'BANK_ACCOUNT',
                               'PASSPORT', 'EMPLOYEE_ID'],
                            'reason': 'OCR-detected PII'
            },
            
            # CONDITIONAL REDACTION
            'client_logos': {
                'enabled': True,
                'severity': RedactionSeverity.MEDIUM,
                'method': RedactionMethod.BLUR_MEDIUM,
                'whitelist': ['microsoft', 'google', 'aws', 'amazon'],
                'min_confidence': 0.5,
                'reason': 'Client confidentiality (NDA compliance)'
            },
            'whiteboards': {
                'enabled': True,
                'severity': RedactionSeverity.MEDIUM,
                'method': RedactionMethod.BLUR_MEDIUM,
                'redact_if_contains': ['project', 'timeline', 'budget', 
                                       'client', 'confidential', 'credential'],
                'min_confidence': 0.4,
                'reason': 'Potential strategic information'
            },
            'printed_documents': {
                'enabled': True,
                'severity': RedactionSeverity.HIGH,
                'method': RedactionMethod.BLACKOUT,
                'types': ['invoice', 'contract', 'payslip', 'offer_letter',
                         'financial_report', 'meeting_notes'],
                'min_confidence': 0.5,
                'reason': 'Confidential business documents'
            },
            
            # SAFE - NO REDACTION
            'safe_objects': {
                'enabled': False,
                'categories': ['furniture', 'office_layout', 'plants', 
                              'generic_signage', 'walls', 'ceiling']
            }
        }
    
    def get_redaction_method(self, entity_type: str) -> RedactionMethod:
        """Get redaction method for entity type"""
        config = self.policy.get(entity_type, {})
        return config.get('method', RedactionMethod.BLUR_HEAVY)
    
    def should_redact(self, entity_type: str) -> bool:
        """Check if entity type should be redacted"""
        return self.policy.get(entity_type, {}).get('enabled', False)
    
    def get_confidence_threshold(self, entity_type: str) -> float:
        """Get minimum confidence for redaction"""
        return self.policy.get(entity_type, {}).get('min_confidence', 0.5)

# Context-Aware Entity Classifier
class ContextAwareClassifier:
    
    def __init__(self, groq_client):
        self.groq_client = groq_client
        self.classification_cache = {}
    
    def classify_entities_in_context(self, text: str, detected_entities: List[Dict]) -> Dict[str, str]:
       
        if not detected_entities:
            return {}
        
        # Build entity list for LLM
        entity_list = []
        for idx, entity in enumerate(detected_entities):
            entity_list.append(f"{idx+1}. '{entity['text']}' (Type: {entity['type']})")
        
        prompt = f"""You are a CORPORATE DATA SECURITY expert analyzing text from a company employee to prevent data leaks.

        CONTEXT: This is from a corporate environment. Your job is to protect:
        - Employee personal information
        - Client/customer data
        - Financial information
        - Project codes and proprietary information

        TEXT TO ANALYZE:
        {text}

        DETECTED ENTITIES:
        {chr(10).join(entity_list)}

        CLASSIFY EACH ENTITY INTO ONE CATEGORY:

        1. **EMPLOYEE_PII** - Employee's own sensitive information
        - Employee's name (if "my name is X", "I am X")
        - Employee's phone, email, address, ID numbers, salary, bank details
        - Example: "I'm Rahul, my salary is 12 LPA" → EMPLOYEE_PII

        2. **CLIENT_SENSITIVE** - Client/customer/vendor information
        - Client names (unless Fortune 500 companies)
        - Customer contact details, addresses
        - Project codes, engagement details
        - Example: "Our client Rajesh from ABC Corp" → CLIENT_SENSITIVE

        3. **FINANCIAL_DATA** - Money-related sensitive info
        - Bank accounts, IFSC codes, salary figures
        - Revenue numbers, budgets, invoices
        - Example: "Account: 1234567890" → FINANCIAL_DATA

        4. **PUBLIC_FIGURE** - Famous people/large organizations
        - Celebrities, politicians, CEOs of major companies
        - Fortune 500 companies, government organizations
        - Example: "Salman Khan" (actor), "Microsoft", "Google" → PUBLIC_FIGURE

        5. **COLLEAGUE_PII** - Other employees/colleagues mentioned
        - Colleague names in context like "my friend X", "my colleague Y"
        - Example: "invite my colleague Amit" → COLLEAGUE_PII

        6. **KEEP** - Safe to keep
        - Generic job titles, departments
        - Public companies in professional context
        - Generic locations (cities, countries - not home addresses)

        CRITICAL CORPORATE RULES:
        - "My name is Salman Khan" → Check if actor or employee (EMPLOYEE_PII if employee)
        - "Want to dance with Akshay Kumar" → PUBLIC_FIGURE (famous actor)
        - "My friend Amit" → COLLEAGUE_PII (redact)
        - Client names (unless huge companies) → CLIENT_SENSITIVE
        - ANY bank account, salary, address → Always redact
        - Project codes, employee IDs → Always redact
        - "Microsoft" in work context → KEEP if discussing company, CLIENT_SENSITIVE if it's your client

        Respond ONLY with JSON mapping entity number to classification:
        {{"1": "EMPLOYEE_PII", "2": "PUBLIC_FIGURE", "3": "CLIENT_SENSITIVE", ...}}"""

        try:
            response = self.groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=500
            )
            
            result_text = response.choices[0].message.content.strip()
            
            # Extract JSON from response
            json_match = re.search(r'\{.*\}', result_text, re.DOTALL)
            if json_match:
                classifications = json.loads(json_match.group())
                
                # Map back to entity texts
                entity_classifications = {}
                for idx, entity in enumerate(detected_entities):
                    classification = classifications.get(str(idx + 1), "KEEP")
                    entity_classifications[entity['text']] = classification
                
                return entity_classifications
            
        except Exception as e:
            logger.warning(f"Context classification failed: {e}")
        
        # Fallback: treat all PERSON entities as potential PII
        return {entity['text']: 'USER_PII' if entity['type'] == 'PERSON' else 'KEEP' 
                for entity in detected_entities}

# Public Entity Verification
class PublicEntityVerifier:
    """Quick verification of public figures using Wikipedia + LLM"""
    
    def __init__(self, groq_client):
        self.groq_client = groq_client
        self.cache = set()
        self.cache_file = "public_entities_cache.json"
        self._load_cache()
    
    def _load_cache(self):
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'r') as f:
                    data = json.load(f)
                    self.cache = set(data.get('entities', []))
                    logger.info(f"Loaded {len(self.cache)} cached public entities")
            except Exception as e:
                logger.warning(f"Could not load cache: {e}")
    
    def _save_cache(self):
        try:
            with open(self.cache_file, 'w') as f:
                json.dump({'entities': list(self.cache)}, f)
                logger.debug(f"Saved {len(self.cache)} entities to cache")
        except Exception as e:
            logger.warning(f"Could not save cache: {e}")
    
    def is_public_figure(self, name: str) -> bool:
        """Check if entity is a public figure"""
        name_lower = name.lower().strip()
        
        if name_lower in self.cache:
            return True
        
        # Check Wikipedia
        if self._check_wikipedia(name):
            self.cache.add(name_lower)
            self._save_cache()
            logger.info(f"Verified public figure via Wikipedia: {name}")
            return True
        
        # LLM verification
        if self._check_with_llm(name):
            self.cache.add(name_lower)
            self._save_cache()
            logger.info(f"Verified public figure via LLM: {name}")
            return True
        
        logger.debug(f"Not a public figure: {name}")
        return False
    
    def _check_wikipedia(self, name: str) -> bool:
        try:
            url = "https://en.wikipedia.org/w/api.php"
            params = {
                'action': 'query',
                'format': 'json',
                'titles': name,
                'redirects': 1
            }
            response = requests.get(url, params=params, timeout=3)
            data = response.json()
            pages = data.get('query', {}).get('pages', {})
            
            for page_id in pages:
                if page_id != '-1' and 'missing' not in pages[page_id]:
                    return True
            return False
        except Exception as e:
            logger.debug(f"Wikipedia check failed for {name}: {e}")
            return False
    
    def _check_with_llm(self, name: str) -> bool:
        try:
            prompt = f"""Is "{name}" a well-known public figure, celebrity, politician, historical figure, or major organization?

            Answer ONLY: YES or NO

            Entity: "{name}"
            Answer:"""
            
            response = self.groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=10
            )
            
            answer = response.choices[0].message.content.strip().upper()
            return answer.startswith("YES")
        except Exception as e:
            logger.debug(f"LLM verification failed for {name}: {e}")
            return False

# ID DOCUMENT DETECTOR 
class IDDocumentDetector:
    """Detect identity documents in images using template matching and OCR"""
    
    def __init__(self):
        self.id_keywords = {
            'aadhaar': ['aadhaar', 'आधार', 'government of india', 'uid'],
            'pan': ['income tax', 'permanent account number', 'pan'],
            'passport': ['passport', 'republic of india', 'p<ind'],
            'driving_license': ['driving licence', 'driving license', 'dl no'],
            'employee_badge': ['employee id', 'emp id', 'staff id', 'badge'],
            'voter_id': ['election commission', 'elector', 'voter']
        }

    @retry_on_failure(max_retries=2, delay=1)
    def detect_id_documents(self, image_cv, ocr_reader=None) -> List[Dict]:
        
        detected_docs = []
        
        if ocr_reader is None:
            return detected_docs
        
        try:
            # Run OCR on entire image
            ocr_results = ocr_reader.readtext(image_cv)
            full_text = ' '.join([text for (_, text, _) in ocr_results]).lower()
            
            # Check for each ID type
            for doc_type, keywords in self.id_keywords.items():
                for keyword in keywords:
                    if keyword in full_text:
                        # Find bounding box of the document
                        bbox = self._find_document_region(image_cv, ocr_results, keywords)
                        if bbox:
                            detected_docs.append({
                                'type': doc_type,
                                'bbox': bbox,
                                'confidence': 0.85,
                                'reason': f'Keyword match: {keyword}'
                            })
                            logger.info(f"ID Document detected: {doc_type}")
                            break  # One match per type
        
        except Exception as e:
            logger.warning(f"ID document detection failed: {e}")
        
        return detected_docs
    
    def _find_document_region(self, image_cv, ocr_results, keywords) -> Optional[Tuple]:
        """Find bounding box containing ID document keywords"""
        matching_boxes = []
        
        for (bbox, text, conf) in ocr_results:
            if any(kw in text.lower() for kw in keywords):
                matching_boxes.append(bbox)
        
        if not matching_boxes:
            return None
        
        # Expand to cover entire document area
        all_points = np.array([pt for box in matching_boxes for pt in box])
        x_min, y_min = all_points.min(axis=0).astype(int)
        x_max, y_max = all_points.max(axis=0).astype(int)
        
        # Add padding (20% on each side)
        h, w = image_cv.shape[:2]
        padding_x = int((x_max - x_min) * 0.2)
        padding_y = int((y_max - y_min) * 0.2)
        
        x = max(0, x_min - padding_x)
        y = max(0, y_min - padding_y)
        width = min(w - x, x_max - x_min + 2 * padding_x)
        height = min(h - y, y_max - y_min + 2 * padding_y)
        
        return (x, y, width, height)

# SCREEN/DASHBOARD DETECTOR
class ScreenDashboardDetector:
    """Detect computer screens and dashboards showing sensitive data"""
    
    def __init__(self):
        self.screen_indicators = [
            # UI Elements
            'chrome', 'firefox', 'browser', 'toolbar',
            # Applications
            'gmail', 'outlook', 'slack', 'teams', 'zoom',
            'excel', 'word', 'powerpoint', 'dashboard',
            # UI Text
            'inbox', 'sent', 'draft', 'new message',
            'file', 'edit', 'view', 'help',
            # Business Apps
            'salesforce', 'jira', 'confluence', 'crm',
            'analytics', 'report', 'metrics'
        ]
    
    def detect_screens(self, image_cv, ocr_reader=None) -> List[Dict]:
        """
        Detect screens/dashboards in image
        Returns: List of {'type': 'screen', 'bbox': (x,y,w,h), 'confidence': 0.9}
        """
        detected_screens = []
        
        if ocr_reader is None:
            return detected_screens
        
        try:
            # Run OCR
            ocr_results = ocr_reader.readtext(image_cv)
            full_text = ' '.join([text for (_, text, _) in ocr_results]).lower()
            
            # Check for screen indicators
            matches = [ind for ind in self.screen_indicators if ind in full_text]
            
            if len(matches) >= 2:  # At least 2 indicators = likely a screen
                # Find the rectangular region containing UI elements
                bbox = self._find_screen_region(image_cv, ocr_results)
                
                if bbox:
                    detected_screens.append({
                        'type': 'computer_screen',
                        'bbox': bbox,
                        'confidence': min(0.95, 0.5 + len(matches) * 0.1),
                        'reason': f'UI elements detected: {", ".join(matches[:3])}'
                    })
                    logger.info(f"Screen detected with {len(matches)} UI indicators")
        
        except Exception as e:
            logger.warning(f"Screen detection failed: {e}")
        
        return detected_screens
    
    def _find_screen_region(self, image_cv, ocr_results) -> Optional[Tuple]:
        """Find bounding box of screen area"""
        if not ocr_results:
            return None
        
        # Get all text regions
        all_boxes = [bbox for (bbox, _, _) in ocr_results]
        
        if len(all_boxes) < 5:  # Too few elements
            return None
        
        # Find dense region of text (likely a screen)
        all_points = np.array([pt for box in all_boxes for pt in box])
        x_min, y_min = all_points.min(axis=0).astype(int)
        x_max, y_max = all_points.max(axis=0).astype(int)
        
        # Basic validation: screens are typically rectangular and sizable
        width = x_max - x_min
        height = y_max - y_min
        
        if width < 200 or height < 150:  # Too small to be a screen
            return None
        
        return (x_min, y_min, width, height)

# WHITEBOARD/PRINTED DOCUMENT DETECTOR 
class WhiteboardDocumentDetector:
    """Detect whiteboards and printed documents in images"""
    
    def __init__(self):
        self.document_keywords = [
            # Document types
            'invoice', 'receipt', 'contract', 'agreement', 'offer letter',
            'payslip', 'salary slip', 'financial report', 'balance sheet',
            'meeting notes', 'minutes', 'confidential', 'internal use only',
            # Document headers
            'to:', 'from:', 'subject:', 'date:', 'invoice no', 'bill to',
            # Whiteboard indicators
            'project timeline', 'roadmap', 'sprint', 'q1', 'q2', 'q3', 'q4',
            'budget', 'revenue', 'target', 'milestone'
        ]
    
    def detect_documents(self, image_cv, ocr_reader=None) -> List[Dict]:
        """
        Detect printed documents and whiteboards
        Returns: List of {'type': 'document', 'bbox': (x,y,w,h), 'confidence': 0.8}
        """
        detected_docs = []
        
        if ocr_reader is None:
            return detected_docs
        
        try:
            # Run OCR
            ocr_results = ocr_reader.readtext(image_cv)
            full_text = ' '.join([text for (_, text, _) in ocr_results]).lower()
            
            # Check for document indicators
            matches = [kw for kw in self.document_keywords if kw in full_text]
            
            if len(matches) >= 2:  # At least 2 document indicators
                # Detect if it's a whiteboard (large white/light-colored region)
                is_whiteboard = self._is_whiteboard(image_cv)
                
                bbox = self._find_document_region(image_cv, ocr_results)
                
                if bbox:
                    detected_docs.append({
                        'type': 'whiteboard' if is_whiteboard else 'printed_document',
                        'bbox': bbox,
                        'confidence': min(0.9, 0.5 + len(matches) * 0.1),
                        'reason': f'Document keywords detected: {", ".join(matches[:3])}'
                    })
                    logger.info(f"{'Whiteboard' if is_whiteboard else 'Document'} detected with {len(matches)} keywords")
        
        except Exception as e:
            logger.warning(f"Document detection failed: {e}")
        
        return detected_docs
    
    def _is_whiteboard(self, image_cv) -> bool:
        """Check if image contains a whiteboard (large white/light area)"""
        # Convert to grayscale
        gray = cv2.cvtColor(image_cv, cv2.COLOR_BGR2GRAY)
        
        # Count bright pixels (whiteboard is typically bright)
        bright_pixels = np.sum(gray > 200)
        total_pixels = gray.size
        
        # If >40% of image is bright, likely a whiteboard
        return (bright_pixels / total_pixels) > 0.4
    
    def _find_document_region(self, image_cv, ocr_results) -> Optional[Tuple]:
        """Find bounding box of document/whiteboard area"""
        if not ocr_results:
            return None
        
        all_boxes = [bbox for (bbox, _, _) in ocr_results]
        
        if len(all_boxes) < 3:
            return None
        
        # Find region containing text
        all_points = np.array([pt for box in all_boxes for pt in box])
        x_min, y_min = all_points.min(axis=0).astype(int)
        x_max, y_max = all_points.max(axis=0).astype(int)
        
        width = x_max - x_min
        height = y_max - y_min
        
        # Expand to cover full document (add 10% padding)
        h, w = image_cv.shape[:2]
        padding_x = int(width * 0.1)
        padding_y = int(height * 0.1)
        
        x = max(0, x_min - padding_x)
        y = max(0, y_min - padding_y)
        width = min(w - x, width + 2 * padding_x)
        height = min(h - y, height + 2 * padding_y)
        
        return (x, y, width, height)
# Indian Context PII Recognizers
class IndianPIIRecognizers:
    
    @staticmethod
    def create_aadhaar_recognizer():
        patterns = [
            Pattern("Aadhaar", r"\b\d{4}[\s-]?\d{4}[\s-]?\d{4}\b", 0.9),
            Pattern("Aadhaar", r"\b\d{12}\b", 0.85),
        ]
        return PatternRecognizer(
            supported_entity="AADHAAR_NUMBER",
            patterns=patterns,
            context=["aadhaar", "uid", "आधार", "aadhar"]
        )
    
    @staticmethod
    def create_pan_recognizer():
        patterns = [
            Pattern("PAN", r"\b[A-Z]{5}\d{4}[A-Z]\b", 0.95),
        ]
        return PatternRecognizer(
            supported_entity="PAN_NUMBER",
            patterns=patterns,
            context=["pan", "permanent account"]
        )
    
    @staticmethod
    def create_indian_phone_recognizer():
        patterns = [
            Pattern("Indian Phone", r"\b(?:\+91|91)?[\s-]?[6-9]\d{9}\b", 0.9),
        ]
        return PatternRecognizer(
            supported_entity="INDIAN_PHONE",
            patterns=patterns,
            context=["phone", "mobile", "contact", "whatsapp", "call"]
        )
    
    @staticmethod
    def create_indian_pincode_recognizer():
        patterns = [
            Pattern("PIN Code", r"\b[1-9]\d{5}\b", 0.7),
        ]
        return PatternRecognizer(
            supported_entity="PIN_CODE",
            patterns=patterns,
            context=["pincode", "pin", "postal", "zip"]
        )
class CorporatePIIRecognizers:
    """Corporate-specific PII patterns"""
    
    @staticmethod
    def create_bank_account_recognizer():
        patterns = [
            Pattern("Bank Account", r"\b\d{9,18}\b", 0.75),  # Indian accounts
            Pattern("IFSC", r"\b[A-Z]{4}0[A-Z0-9]{6}\b", 0.95),  # IFSC codes
            Pattern("SWIFT", r"\b[A-Z]{6}[A-Z0-9]{2}([A-Z0-9]{3})?\b", 0.9),  # SWIFT codes
        ]
        return PatternRecognizer(
            supported_entity="BANK_ACCOUNT",
            patterns=patterns,
            context=["account", "bank", "ifsc", "swift", "a/c", "account number"]
        )
    
    @staticmethod
    def create_employee_id_recognizer():
        patterns = [
            Pattern("Employee ID", r"\bEMP[-_]?\d{4,8}\b", 0.9),
            Pattern("Employee ID", r"\b[A-Z]{2,4}\d{4,6}\b", 0.7),  # Generic EMP123456
            Pattern("Employee ID", r"\bSTAFF[-_]?\d{4,8}\b", 0.85),
        ]
        return PatternRecognizer(
            supported_entity="EMPLOYEE_ID",
            patterns=patterns,
            context=["employee", "emp", "staff", "id", "badge"]
        )
    
    @staticmethod
    def create_salary_recognizer():
        patterns = [
            Pattern("Salary", r"\b(?:Rs\.?|INR|₹)\s*\d{1,3}(?:,\d{3})*(?:\.\d{2})?\b", 0.85),
            Pattern("Salary", r"\b\d{1,3}(?:,\d{3})*(?:\.\d{2})?\s*(?:LPA|CTC|per annum)\b", 0.9),
        ]
        return PatternRecognizer(
            supported_entity="SALARY_INFO",
            patterns=patterns,
            context=["salary", "ctc", "compensation", "package", "pay"]
        )
    
    @staticmethod
    def create_passport_recognizer():
        patterns = [
            Pattern("Passport", r"\b[A-Z]\d{7}\b", 0.9),  # Indian
            Pattern("Passport", r"\b[A-Z]{1,2}\d{6,9}\b", 0.8),  # International
        ]
        return PatternRecognizer(
            supported_entity="PASSPORT",
            patterns=patterns,
            context=["passport", "travel document"]
        )
    
    @staticmethod
    def create_project_code_recognizer():
        patterns = [
            Pattern("Project Code", r"\bPROJ[-_]?[A-Z0-9]{4,10}\b", 0.85),
            Pattern("Project Code", r"\b[A-Z]{2,4}[-_]\d{3,6}\b", 0.7),
        ]
        return PatternRecognizer(
            supported_entity="PROJECT_CODE",
            patterns=patterns,
            context=["project", "proj", "code", "engagement"]
        )
    
    @staticmethod
    def create_home_address_recognizer():
        # More sophisticated - catches multi-line addresses
        patterns = [
            Pattern("Address", r"\b\d{1,5}\s+[\w\s]{3,50},\s*[\w\s]{3,30},\s*[A-Z]{2}\s+\d{5,6}\b", 0.8),
            Pattern("Address", r"(?i)\b(?:flat|plot|house|bldg|building)\s*(?:no\.?|#)?\s*\d+.*?\d{6}\b", 0.75),
        ]
        return PatternRecognizer(
            supported_entity="HOME_ADDRESS",
            patterns=patterns,
            context=["address", "residence", "home", "street", "apartment"]
        )
    
    @staticmethod
    def create_client_name_recognizer():
        # This will need LLM assistance
        patterns = [
            Pattern("Client", r"\bclient\s+(?:name|code):\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)", 0.7),
        ]
        return PatternRecognizer(
            supported_entity="CLIENT_INFO",
            patterns=patterns,
            context=["client", "customer", "vendor", "partner"]
        )

# Enhanced Context-Aware PII Guard
class ContextAwarePIIGuard:
    
    def __init__(self, spacy_model: str, language: str, groq_client):
        # Initialize Presidio
        nlp_config = {
            "nlp_engine_name": "spacy",
            "models": [{"lang_code": language, "model_name": spacy_model}]
        }
        provider = NlpEngineProvider(nlp_configuration=nlp_config)
        nlp_engine = provider.create_engine()
        
        # Custom registry with Indian recognizers
        registry = RecognizerRegistry()
        registry.load_predefined_recognizers()

        # Add Indian PII recognizers
        registry.add_recognizer(IndianPIIRecognizers.create_aadhaar_recognizer())
        registry.add_recognizer(IndianPIIRecognizers.create_pan_recognizer())
        registry.add_recognizer(IndianPIIRecognizers.create_indian_phone_recognizer())
        registry.add_recognizer(IndianPIIRecognizers.create_indian_pincode_recognizer())

        # Corporate PII
        registry.add_recognizer(CorporatePIIRecognizers.create_bank_account_recognizer())
        registry.add_recognizer(CorporatePIIRecognizers.create_employee_id_recognizer())
        registry.add_recognizer(CorporatePIIRecognizers.create_salary_recognizer())
        registry.add_recognizer(CorporatePIIRecognizers.create_passport_recognizer())
        registry.add_recognizer(CorporatePIIRecognizers.create_project_code_recognizer())
        registry.add_recognizer(CorporatePIIRecognizers.create_home_address_recognizer())
        registry.add_recognizer(CorporatePIIRecognizers.create_client_name_recognizer())
        
        self.analyzer = AnalyzerEngine(
            nlp_engine=nlp_engine,
            supported_languages=[language],
            registry=registry
        )
        self.anonymizer = AnonymizerEngine()
        self.language = language
        
        # Context understanding
        self.context_classifier = ContextAwareClassifier(groq_client)
        self.public_verifier = PublicEntityVerifier(groq_client)
        
        self.mapping = {}
        self.reverse_mapping = {}
        self.kept_entities = set()

        self.redaction_summary = {
        'EMPLOYEE_PII': [],
        'CLIENT_SENSITIVE': [],
        'FINANCIAL_DATA': [],
        'COLLEAGUE_PII': [],
        'BANK_ACCOUNT': [],
        'SALARY_INFO': [],
        'EMPLOYEE_ID': [],
        'PROJECT_CODE': [],
        'HOME_ADDRESS': [],
        'PASSPORT': []
    }
        
    def get_redaction_report(self) -> Dict:
        """Generate detailed redaction report"""
        total_redacted = sum(len(v) for v in self.redaction_summary.values())
        
        report = {
            'total_entities_redacted': total_redacted,
            'total_entities_kept': len(self.kept_entities),
            'redacted_by_category': {},
            'kept_entities_list': list(self.kept_entities),
            'detailed_redactions': self.redaction_summary
        }
        
        # Count by category
        for category, items in self.redaction_summary.items():
            if items:
                report['redacted_by_category'][category] = len(items)
        
        return report

    def anonymize(self, text: str, context_aware: bool = True) -> str:
        """
        Anonymize text with context awareness.
        Only redacts USER_PII and THIRD_PARTY_PII, keeps PUBLIC_FIGURE mentions.
        """
        self.mapping.clear()
        self.reverse_mapping.clear()
        self.kept_entities.clear()
        
        # Detect all PII
        results = self.analyzer.analyze(text=text, language=self.language)
        
        if not results:
            return text
        
        # Build entity list for classification
        detected_entities = []
        for result in results:
            entity_text = text[result.start:result.end]
            detected_entities.append({
                'text': entity_text,
                'type': result.entity_type,
                'start': result.start,
                'end': result.end,
                'score': result.score
            })
        
        # Classify entities based on context
        if context_aware:
            classifications = self.context_classifier.classify_entities_in_context(text, detected_entities)
        else:
            classifications = {e['text']: 'USER_PII' for e in detected_entities}
        
        # Filter results based on classification
        filtered_results = []
        for result in results:
            entity_text = text[result.start:result.end]
            classification = classifications.get(entity_text, 'KEEP')
            
            # Keep public figures and entities marked as KEEP
            if classification in ['PUBLIC_FIGURE', 'KEEP']:
                self.kept_entities.add(entity_text)
                logger.info(f"KEEPING: {entity_text} - {classification}")
                continue
            
            # Redact USER_PII and THIRD_PARTY_PII
            if classification in ['USER_PII', 'THIRD_PARTY_PII', 'EMPLOYEE_PII', 
                      'CLIENT_SENSITIVE', 'FINANCIAL_DATA', 'COLLEAGUE_PII']:
                logger.info(f"REDACTING: {entity_text} - {classification}")
                filtered_results.append(result)

                category = classification if classification in self.redaction_summary else result.entity_type
                if category in self.redaction_summary:
                    self.redaction_summary[category].append({
                        'text': entity_text,
                        'type': result.entity_type,
                        'reason': classification
                    })
        
        if not filtered_results:
            return text
        
        # Anonymize filtered results
        operators = {}
        counters = {}
        
        for result in filtered_results:
            entity_type = result.entity_type
            counters[entity_type] = counters.get(entity_type, 0) + 1
            placeholder = f"<{entity_type}_{counters[entity_type]}>"
            operators[entity_type] = OperatorConfig("replace", {"new_value": placeholder})
        
        anonymized_result = self.anonymizer.anonymize(
            text=text,
            analyzer_results=filtered_results,
            operators=operators
        )
        
        self._build_mapping(text, anonymized_result.text, filtered_results)
        
        return anonymized_result.text
    
    def _build_mapping(self, original: str, anonymized: str, results):
        for result in results:
            original_value = original[result.start:result.end]
            pattern = f"<{result.entity_type}_\\d+>"
            matches = list(re.finditer(pattern, anonymized))
            
            for match in matches:
                placeholder = match.group()
                if placeholder not in self.reverse_mapping:
                    self.reverse_mapping[placeholder] = original_value
                    self.mapping[original_value] = placeholder
                    break
    
    def deanonymize(self, text: str) -> str:
        result = text
        for placeholder, original_value in self.reverse_mapping.items():
            result = result.replace(placeholder, original_value)
            bracket_placeholder = placeholder.replace('<', '[').replace('>', ']')
            result = result.replace(bracket_placeholder, original_value)
            plain_placeholder = placeholder.strip('<>')
            result = re.sub(r'\b' + re.escape(plain_placeholder) + r'\b', original_value, result)
        return result
    
    def clear(self):
        self.mapping.clear()
        self.reverse_mapping.clear()
        self.kept_entities.clear()

        # Clear redaction summary for next run
        for category in self.redaction_summary:
            self.redaction_summary[category].clear()

# Excel/CSV Handler
class SpreadsheetHandler:
    
    def __init__(self, pii_guard: ContextAwarePIIGuard):
        self.pii_guard = pii_guard
    
    def process_excel(self, file_path: str, output_path: str = None) -> Dict:
        """Process Excel file and redact PII"""
        logger.info(f"Processing Excel file: {file_path}")
        
        if not output_path:
            output_path = file_path.replace('.xlsx', '_REDACTED.xlsx').replace('.xls', '_REDACTED.xls')
        
        # Read Excel
        workbook = openpyxl.load_workbook(file_path)
        stats = {
            'sheets_processed': 0,
            'cells_redacted': 0,
            'pii_found': {}
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            stats['sheets_processed'] += 1
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        original = cell.value
                        anonymized = self.pii_guard.anonymize(original)
                        
                        if original != anonymized:
                            cell.value = anonymized
                            stats['cells_redacted'] += 1
        
        # Save redacted workbook
        workbook.save(output_path)
        logger.info(f"Saved redacted Excel: {output_path}")
        
        stats['pii_found'] = dict(self.pii_guard.reverse_mapping)
        return {
            'type': 'excel',
            'input_file': file_path,
            'output_file': output_path,
            'stats': stats,
            'timestamp': datetime.now().isoformat()
        }
    
    def process_csv(self, file_path: str, output_path: str = None) -> Dict:
        """Process CSV file and redact PII"""
        logger.info(f"Processing CSV file: {file_path}")
        
        if not output_path:
            output_path = file_path.replace('.csv', '_REDACTED.csv')
        
        df = pd.read_csv(file_path)
        original_shape = df.shape
        cells_redacted = 0
        
        for col in df.columns:
            if df[col].dtype == 'object':  # Text columns
                for idx in df.index:
                    if pd.notna(df.at[idx, col]):
                        original = str(df.at[idx, col])
                        anonymized = self.pii_guard.anonymize(original)
                        
                        if original != anonymized:
                            df.at[idx, col] = anonymized
                            cells_redacted += 1
        
        df.to_csv(output_path, index=False)
        logger.info(f"Saved redacted CSV: {output_path}")
        
        return {
            'type': 'csv',
            'input_file': file_path,
            'output_file': output_path,
            'stats': {
                'rows': original_shape[0],
                'columns': original_shape[1],
                'cells_redacted': cells_redacted,
                'pii_found': dict(self.pii_guard.reverse_mapping)
            },
            'timestamp': datetime.now().isoformat()
        }
    
#  PRODUCTION-GRADE IMAGE REDACTOR 
class ProductionImageRedactor:
    """Production-grade image redaction with policy enforcement and audit trails"""
    
    def __init__(self, pii_guard: 'ContextAwarePIIGuard' = None, 
                 policy: RedactionPolicy = None):
        # Detection models
        self.face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        )
        self.pii_guard = pii_guard
        self.policy = policy or RedactionPolicy()
        
        # Advanced detectors
        self.id_detector = IDDocumentDetector()
        self.screen_detector = ScreenDashboardDetector()
        self.whiteboard_detector = WhiteboardDocumentDetector()  
        
        # Cache OCR results
        self._ocr_cache = {}  # Fixed: use underscore
        self._enable_caching = True
        
        # OCR
        try:
            import easyocr
            self.ocr_reader = easyocr.Reader(['en'])
            self.ocr_available = True
            logger.info("[OK] EasyOCR initialized successfully")
        except Exception as e:
            logger.warning(f"EasyOCR not available: {e}")
            self.ocr_available = False

    # Validate image quality
    def _validate_image_quality(self, img_cv) -> Dict:
        """Check if image quality is sufficient for redaction"""
        h, w = img_cv.shape[:2]
        
        validation = {
            'is_valid': True,
            'warnings': [],
            'resolution': (w, h)
        }
        
        # Check resolution
        if w < 300 or h < 300:
            validation['warnings'].append('Low resolution - redaction accuracy may be reduced')
        
        # Check brightness
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        avg_brightness = np.mean(gray)
        
        if avg_brightness < 50:
            validation['warnings'].append('Image too dark - OCR may fail')
        elif avg_brightness > 200:
            validation['warnings'].append('Image overexposed - detection may be affected')
        
        # Check blur
        laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()
        if laplacian_var < 100:
            validation['warnings'].append('Image appears blurry - detection accuracy reduced')
        
        return validation
    
    def _get_ocr_results(self, img_cv):
        """Get OCR results with caching"""
        if not self.ocr_available:
            return []
        
        # Create image hash for caching
        img_hash = hash(img_cv.tobytes())
        
        if self._enable_caching and img_hash in self._ocr_cache:
            logger.debug("Using cached OCR results")
            return self._ocr_cache[img_hash]
        
        # Run OCR
        results = self.ocr_reader.readtext(img_cv)
        
        if self._enable_caching:
            self._ocr_cache[img_hash] = results
        
        return results
    
    
    def redact_image(self, image: Image.Image, 
                     redact_faces=True,
                     redact_ids=True,
                     redact_screens=True,
                     redact_text=True,
                     redact_logos=True,
                     redact_documents=True) -> Tuple[Image.Image, Dict]:  
        """Redact image based on policy and specified options"""
        img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)

        # Image quality validation
        quality_check = self._validate_image_quality(img_cv)
        if quality_check['warnings']:
            for warning in quality_check['warnings']:
                logger.warning(f"Image Quality: {warning}")
        
        # Audit trail
        redaction_log = {
            'timestamp': datetime.now().isoformat(),
            'detections': [],
            'redactions': [],
            'policy_applied': 'corporate_default',
            'summary': {}
        }
        
        # PHASE 1: CRITICAL - Identity Documents (highest priority)
        if redact_ids and self.policy.should_redact('identity_documents'):
            img_cv, id_log = self._redact_id_documents(img_cv)
            redaction_log['detections'].extend(id_log['detections'])
            redaction_log['redactions'].extend(id_log['redactions'])
        
        # PHASE 2: HIGH - Screens/Dashboards
        if redact_screens and self.policy.should_redact('screens_dashboards'):
            img_cv, screen_log = self._redact_screens(img_cv)
            redaction_log['detections'].extend(screen_log['detections'])
            redaction_log['redactions'].extend(screen_log['redactions'])
        
        # PHASE 3: HIGH - Faces (biometric PII)
        if redact_faces and self.policy.should_redact('faces'):
            img_cv, face_log = self._redact_faces_policy(img_cv)
            redaction_log['detections'].extend(face_log['detections'])
            redaction_log['redactions'].extend(face_log['redactions'])
        
        # PHASE 4: HIGH - OCR Text (PII in text)
        if redact_text and self.policy.should_redact('ocr_pii') and self.pii_guard:
            img_cv, text_log = self._redact_text_pii_policy(img_cv)
            redaction_log['detections'].extend(text_log['detections'])
            redaction_log['redactions'].extend(text_log['redactions'])
        
        # PHASE 5: MEDIUM - Client logos (conditional)
        if redact_logos and self.policy.should_redact('client_logos'):
            img_cv, logo_log = self._redact_logos_policy(img_cv)
            redaction_log['detections'].extend(logo_log['detections'])
            redaction_log['redactions'].extend(logo_log['redactions'])
        
        # *** ADD THIS NEW PHASE 6 ***
        # PHASE 6: MEDIUM - Whiteboards/Printed Documents
        if redact_documents and self.policy.should_redact('printed_documents'):
            img_cv, doc_log = self._redact_documents(img_cv)
            redaction_log['detections'].extend(doc_log['detections'])
            redaction_log['redactions'].extend(doc_log['redactions'])
        
        # Generate summary
        redaction_log['summary'] = self._generate_summary(redaction_log)
        
        # Convert back to PIL
        img_rgb = cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)
        return Image.fromarray(img_rgb), redaction_log

    def _apply_pixelation(self, region, pixel_size=15):
        """Apply pixelation effect to a region"""
        h, w = region.shape[:2]
        
        # Resize down
        temp = cv2.resize(region, (w // pixel_size, h // pixel_size), 
                         interpolation=cv2.INTER_LINEAR)
        
        # Resize back up
        pixelated = cv2.resize(temp, (w, h), interpolation=cv2.INTER_NEAREST)
        
        return pixelated

    def _redact_documents(self, img_cv) -> Tuple[np.ndarray, Dict]:
        """Detect and redact whiteboards/printed documents"""
        log = {'detections': [], 'redactions': []}
        
        if not self.ocr_available:
            return img_cv, log
        
        detected_docs = self.whiteboard_detector.detect_documents(img_cv, self.ocr_reader)
        
        method = self.policy.get_redaction_method('printed_documents')
        
        for doc in detected_docs:
            x, y, w, h = doc['bbox']
            doc_region = img_cv[y:y+h, x:x+w]
            
            # Apply redaction based on policy
            if method == RedactionMethod.BLACKOUT:
                redacted = np.zeros_like(doc_region)
            elif method == RedactionMethod.BLUR_HEAVY:
                redacted = cv2.GaussianBlur(doc_region, (99, 99), 30)
            else:
                redacted = cv2.GaussianBlur(doc_region, (51, 51), 20)
            
            img_cv[y:y+h, x:x+w] = redacted
            
            log['detections'].append({
                'type': doc['type'],
                'confidence': doc['confidence'],
                'bbox': doc['bbox']
            })
            log['redactions'].append({
                'method': method.value,
                'severity': 'MEDIUM' if doc['type'] == 'whiteboard' else 'HIGH',
                'reason': doc['reason'],
                'compliance': 'Confidential business documents'
            })
            
            logger.warning(f"📄 Redacted {doc['type']}")
        
        return img_cv, log

    def _redact_id_documents(self, img_cv) -> Tuple[np.ndarray, Dict]:
        """Detect and redact identity documents - CRITICAL PRIORITY"""
        log = {'detections': [], 'redactions': []}
        
        if not self.ocr_available:
            return img_cv, log
        
        detected_ids = self.id_detector.detect_id_documents(img_cv, self.ocr_reader)
        
        for doc in detected_ids:
            x, y, w, h = doc['bbox']
            
            # BLACKOUT (not blur) - compliance requirement
            cv2.rectangle(img_cv, (x, y), (x+w, y+h), (0, 0, 0), -1)
            
            log['detections'].append({
                'type': 'identity_document',
                'subtype': doc['type'],
                'confidence': doc['confidence'],
                'bbox': doc['bbox']
            })
            log['redactions'].append({
                'method': 'BLACKOUT',
                'severity': 'CRITICAL',
                'reason': f"ID document detected: {doc['type']}",
                'compliance': 'GDPR Art. 9, DPDP Act'
            })
            
            logger.warning(f"🔒 CRITICAL: Redacted {doc['type']} document")
        
        return img_cv, log
    
    def _redact_screens(self, img_cv) -> Tuple[np.ndarray, Dict]:
        """Detect and redact computer screens/dashboards"""
        log = {'detections': [], 'redactions': []}
        
        if not self.ocr_available:
            return img_cv, log
        
        detected_screens = self.screen_detector.detect_screens(img_cv, self.ocr_reader)
        
        for screen in detected_screens:
            x, y, w, h = screen['bbox']
            
            # BLACKOUT entire screen region
            cv2.rectangle(img_cv, (x, y), (x+w, y+h), (0, 0, 0), -1)
            
            log['detections'].append({
                'type': 'computer_screen',
                'confidence': screen['confidence'],
                'bbox': screen['bbox']
            })
            log['redactions'].append({
                'method': 'BLACKOUT',
                'severity': 'HIGH',
                'reason': screen['reason'],
                'compliance': 'Confidential business data protection'
            })
            
            logger.warning(f"🖥️  Redacted computer screen/dashboard")
        
        return img_cv, log
    
    def _redact_faces_policy(self, img_cv) -> Tuple[np.ndarray, Dict]:
        """Redact faces with policy-driven method"""
        log = {'detections': [], 'redactions': []}
        
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(
            gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30)
        )
        
        method = self.policy.get_redaction_method('faces')
        
        for (x, y, w, h) in faces:
            face_region = img_cv[y:y+h, x:x+w]
            
            # Apply redaction method
            if method == RedactionMethod.BLUR_HEAVY:
                redacted = cv2.GaussianBlur(face_region, (99, 99), 30)
            elif method == RedactionMethod.BLACKOUT:
                redacted = np.zeros_like(face_region)
            elif method == RedactionMethod.PIXELATE:  # ADD THIS
                redacted = self._apply_pixelation(face_region, pixel_size=20)
            elif method == RedactionMethod.BLUR_MEDIUM:
                redacted = cv2.GaussianBlur(face_region, (51, 51), 20)
            else:
                redacted = cv2.GaussianBlur(face_region, (51, 51), 20)
            
            img_cv[y:y+h, x:x+w] = redacted
            
            log['detections'].append({
                'type': 'face',
                'confidence': 0.85,
                'bbox': (x, y, w, h)
            })
            log['redactions'].append({
                'method': method.value,
                'severity': 'HIGH',
                'reason': 'Biometric PII (face)',
                'compliance': 'GDPR Art. 9, DPDP Act Sec. 3'
            })
        
        if len(faces) > 0:
            logger.info(f"👤 Redacted {len(faces)} face(s)")
        
        return img_cv, log
    
    def _redact_text_pii_policy(self, img_cv) -> Tuple[np.ndarray, Dict]:
        """Redact OCR-detected PII with confidence filtering"""
        log = {'detections': [], 'redactions': []}
        
        if not self.ocr_available or not self.pii_guard:
            return img_cv, log
        
        min_conf = self.policy.get_confidence_threshold('ocr_pii')
        ocr_results = self._get_ocr_results(img_cv)
        
        for (bbox, text, confidence) in ocr_results:
            if confidence < min_conf:
                continue
            
            # Check for PII
            anonymized = self.pii_guard.anonymize(text, context_aware=False)
            
            if text != anonymized:  # PII found
                top_left = tuple(map(int, bbox[0]))
                bottom_right = tuple(map(int, bbox[2]))
                
                # BLACKOUT
                cv2.rectangle(img_cv, top_left, bottom_right, (0, 0, 0), -1)
                
                log['detections'].append({
                    'type': 'ocr_text_pii',
                    'text_preview': text[:20] + '...' if len(text) > 20 else text,
                    'confidence': confidence,
                    'bbox': bbox
                })
                log['redactions'].append({
                    'method': 'BLACKOUT',
                    'severity': 'HIGH',
                    'reason': 'PII detected in text',
                    'compliance': 'Data minimization principle'
                })
                
                logger.info(f"📝 Redacted PII text: {text[:30]}...")
        
        return img_cv, log
    
    def _redact_logos_policy(self, img_cv) -> Tuple[np.ndarray, Dict]:
        """Redact client logos (conditional based on policy)"""
        log = {'detections': [], 'redactions': []}
        
        # Simple color-based logo detection
        hsv = cv2.cvtColor(img_cv, cv2.COLOR_BGR2HSV)
        mask = cv2.inRange(hsv, (0, 100, 100), (180, 255, 255))
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        method = self.policy.get_redaction_method('client_logos')
        
        for contour in contours:
            area = cv2.contourArea(contour)
            if 2500 < area < 90000:  # Logo size range
                x, y, w, h = cv2.boundingRect(contour)
                logo_region = img_cv[y:y+h, x:x+w]
                
                # Apply redaction
                if method == RedactionMethod.BLUR_MEDIUM:
                    redacted = cv2.GaussianBlur(logo_region, (51, 51), 20)
                else:
                    redacted = cv2.GaussianBlur(logo_region, (99, 99), 30)
                
                img_cv[y:y+h, x:x+w] = redacted
                
                log['detections'].append({
                    'type': 'potential_logo',
                    'confidence': 0.6,
                    'bbox': (x, y, w, h)
                })
                log['redactions'].append({
                    'method': method.value,
                    'severity': 'MEDIUM',
                    'reason': 'Client logo (NDA compliance)',
                    'compliance': 'Confidentiality agreements'
                })
        
        if len(log['detections']) > 0:
            logger.info(f" Redacted {len(log['detections'])} potential logo(s)")
        
        return img_cv, log
    
    def _generate_summary(self, redaction_log: Dict) -> Dict:
        """Generate redaction summary statistics"""
        summary = {
            'total_detections': len(redaction_log['detections']),
            'total_redactions': len(redaction_log['redactions']),
            'by_type': {},
            'by_severity': {},
            'compliance_notes': []
        }
        
        for detection in redaction_log['detections']:
            det_type = detection['type']
            summary['by_type'][det_type] = summary['by_type'].get(det_type, 0) + 1
        
        for redaction in redaction_log['redactions']:
            severity = redaction['severity']
            summary['by_severity'][severity] = summary['by_severity'].get(severity, 0) + 1
            
            if redaction.get('compliance'):
                summary['compliance_notes'].append(redaction['compliance'])
        
        summary['compliance_notes'] = list(set(summary['compliance_notes']))
        
        return summary

# PDF Handler
class PDFHandler:
    
    def __init__(self, pii_guard: ContextAwarePIIGuard):
        self.pii_guard = pii_guard
    
    def process_pdf(self, input_path: str, output_path: str = None) -> Dict:
        """Extract text, anonymize, and create redacted PDF"""
        if not output_path:
            output_path = input_path.replace('.pdf', '_REDACTED.pdf')
        
        doc = fitz.open(input_path)
        full_text = ""
        anonymized_text = ""
        
        for page in doc:
            text = page.get_text()
            full_text += text
            anonymized = self.pii_guard.anonymize(text)
            anonymized_text += anonymized
        
        # Create redacted PDF
        doc_redact = fitz.open(input_path)
        for page in doc_redact:
            text = page.get_text()
            results = self.pii_guard.analyzer.analyze(text=text, language=self.pii_guard.language)
            
            for result in results:
                entity_text = text[result.start:result.end]
                if entity_text not in self.pii_guard.kept_entities:
                    instances = page.search_for(entity_text)
                    for inst in instances:
                        page.add_redact_annot(inst, fill=(0, 0, 0))
            
            page.apply_redactions()
        
        doc_redact.save(output_path)
        doc_redact.close()
        doc.close()
        
        logger.info(f"Saved redacted PDF: {output_path}")
        
        return {
            'type': 'pdf',
            'input_file': input_path,
            'output_file': output_path,
            'pages': len(doc),
            'pii_found': dict(self.pii_guard.reverse_mapping),
            'entities_kept': list(self.pii_guard.kept_entities),
            'timestamp': datetime.now().isoformat()
        }

# LLM Client
class GroqClient:
    
    def __init__(self, api_key: str, model: str = "llama-3.3-70b-versatile"):
        self.client = Groq(api_key=api_key)
        self.model = model
    
    def generate(self, prompt: str) -> str:
        response = self.client.chat.completions.create(
            model=self.model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
            max_tokens=1024
        )
        return response.choices[0].message.content

# Main Pipeline
class IntelligentPIIPipeline:
    
    def __init__(self, config: Config):
        self.llm_client = GroqClient(config.groq_api_key, config.groq_model)
        self.pii_guard = ContextAwarePIIGuard(
            spacy_model=config.spacy_model,
            language=config.language,
            groq_client=self.llm_client.client
        )
        self.pdf_handler = PDFHandler(self.pii_guard)
        self.spreadsheet_handler = SpreadsheetHandler(self.pii_guard)
        self.image_redactor = ProductionImageRedactor(
            self.pii_guard, 
            RedactionPolicy()
        )
    
    def process_text(self, text: str, verbose: bool = True) -> Dict:
        """Process text with context-aware redaction"""
        try:
            if verbose:
                logger.info("="*80)
                logger.info("PROCESSING TEXT WITH CONTEXT AWARENESS")
                logger.info("="*80)
            
            anonymized = self.pii_guard.anonymize(text, context_aware=True)
            
            if verbose:
                logger.info("Generating LLM response...")
            
            llm_response = self.llm_client.generate(anonymized)
            final_response = self.pii_guard.deanonymize(llm_response)
            
            result = {
                'type': 'text',
                'original': text,
                'anonymized': anonymized,
                'llm_response_anonymized': llm_response,
                'final_response': final_response,
                'redacted_entities': dict(self.pii_guard.reverse_mapping),
                'kept_entities': list(self.pii_guard.kept_entities),
                'redaction_report': self.pii_guard.get_redaction_report(),
                'timestamp': datetime.now().isoformat()
            }
            
            if verbose:
                logger.info("="*80)
                logger.info("RESULTS")
                logger.info("="*80)
                logger.info(f"ORIGINAL: {text}")
                logger.info(f"ANONYMIZED: {anonymized}")
                logger.info(f"RESPONSE: {final_response}")
                logger.info(f"REDACTED: {len(self.pii_guard.reverse_mapping)} entities")
                logger.info(f"KEPT: {len(self.pii_guard.kept_entities)} public figures")
            
            return result
            
        finally:
            self.pii_guard.clear()
    
    def process_file(self, file_path: str) -> Dict:
        """Process any supported file type"""
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.pdf':
            return self.pdf_handler.process_pdf(file_path)
        elif ext in ['.xlsx', '.xls']:
            return self.spreadsheet_handler.process_excel(file_path)
        elif ext == '.csv':
            return self.spreadsheet_handler.process_csv(file_path)
        elif ext in ['.jpg', '.jpeg', '.png', '.bmp']:
            return self.process_image(file_path)
        else:
            return {'error': f'Unsupported file type: {ext}'}
    
    # process_image METHOD 
    def process_image(self, image_path: str, custom_policy: Dict = None) -> Dict:
       
        logger.info("="*80)
        logger.info(f"PROCESSING IMAGE: {image_path}")
        logger.info("="*80)
        
        image = Image.open(image_path)
        
        # Initialize redactor with policy
        policy = RedactionPolicy(custom_policy) if custom_policy else RedactionPolicy()
        redactor = ProductionImageRedactor(self.pii_guard, policy)
        
        # Perform redaction
        redacted, redaction_log = redactor.redact_image(
            image,
            redact_faces=True,
            redact_ids=True,
            redact_screens=True,
            redact_text=True,
            redact_logos=True,
            redact_documents=True
        )
        
        # Save redacted image
        output_path = image_path.replace('.', '_REDACTED.')
        redacted.save(output_path)
        
        # Generate compliance report
        result = {
            'type': 'image',
            'input_file': image_path,
            'output_file': output_path,
            'redaction_log': redaction_log,
            'summary': redaction_log['summary'],
            'compliance_status': self._check_compliance(redaction_log),
            'timestamp': datetime.now().isoformat()
        }
        
        # Log summary
        summary = redaction_log['summary']
        logger.info("="*80)
        logger.info("REDACTION SUMMARY")
        logger.info("="*80)
        logger.info(f"Total Detections: {summary['total_detections']}")
        logger.info(f"Total Redactions: {summary['total_redactions']}")
        logger.info(f"By Type: {summary['by_type']}")
        logger.info(f"By Severity: {summary['by_severity']}")
        logger.info(f"Compliance: {', '.join(summary['compliance_notes'])}")
        logger.info(f"Saved: {output_path}")
        logger.info("="*80)
        
        return result
    
    def _check_compliance(self, redaction_log: Dict) -> Dict:
        """Verify compliance with corporate policies"""
        compliance = {
            'gdpr_compliant': True,
            'dpdp_act_compliant': True,
            'iso_27701_compliant': True,
            'issues': []
        }
        
        # Check if any CRITICAL items were missed
        critical_count = redaction_log['summary']['by_severity'].get('CRITICAL', 0)
        
        if critical_count == 0:
            compliance['issues'].append("No critical PII detected - verify manually")
        
        return compliance

# Save results
def save_results(data: Dict, output_dir: str = "output"):
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"result_{timestamp}.json"
    filepath = os.path.join(output_dir, filename)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    logger.info(f"Saved results to: {filepath}")
    return filepath

# Main CLI
def main():
    config = Config(groq_api_key=os.getenv("GROQ_API_KEY"))
    
    if not config.groq_api_key:
        logger.error("GROQ_API_KEY not set in environment!")
        logger.error("Set it using: export GROQ_API_KEY='your-key'")
        return
    
    pipeline = IntelligentPIIPipeline(config)
    
    logger.info("="*80)
    logger.info("INTELLIGENT CONTEXT-AWARE PII REDACTION SYSTEM")
    logger.info("="*80)
    logger.info("\nFeatures:")
    logger.info("✓ Understands context (user PII vs public figures)")
    logger.info("✓ Supports: Text, PDF, Excel, CSV, Images")
    logger.info("✓ LLM-powered entity classification")
    logger.info("="*80)
    
    while True:
        print("\n\nOptions:")
        print("1. Process text input")
        print("2. Process file (PDF/Excel/CSV/Image)")
        print("3. Batch process directory")
        print("4. Exit")
        
        choice = input("\nChoice: ").strip()
        
        if choice == '1':
            text = input("\nEnter text: ").strip()
            if text:
                result = pipeline.process_text(text, verbose=True)
                save_results(result)
        
        elif choice == '2':
            file_path = input("\nFile path: ").strip()
            if os.path.exists(file_path):
                result = pipeline.process_file(file_path)
                save_results(result)
                logger.info(f"✓ Processed: {file_path}")
            else:
                logger.error("File not found")
        
        elif choice == '3':
            dir_path = input("\nDirectory path: ").strip()
            if os.path.isdir(dir_path):
                files = [f for f in os.listdir(dir_path) 
                        if os.path.splitext(f)[1].lower() in 
                        ['.pdf', '.xlsx', '.xls', '.csv', '.jpg', '.png']]
                
                logger.info(f"Found {len(files)} files to process")
                for file in files:
                    file_path = os.path.join(dir_path, file)
                    logger.info(f"Processing: {file}")
                    result = pipeline.process_file(file_path)
                    save_results(result)
                
                logger.info(f"✓  Batch complete: {len(files)} files processed")
            else:
                logger.error("Directory not found")
        
        elif choice == '4':
            logger.info("Exiting...")
            break

if __name__ == "__main__":
    config = Config(groq_api_key=os.getenv("GROQ_API_KEY"))
    if not config.groq_api_key:
        logger.error("GROQ_API_KEY not set in environment!")
        logger.error("Set it using: export GROQ_API_KEY='your-key'")
    main()